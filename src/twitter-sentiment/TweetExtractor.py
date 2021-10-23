import openpyxl
from tweepy import *

from TweetQuery import TweetQuery 
from TweetQueryHandler import TweetQueryHandler
from TweetContainer import TweetContainer
class TweetExtractor:

    def __init__(self, bearer_token: str, file_path: str) -> None:
        self.file_path = file_path
        self.twitter_client = Client(bearer_token=bearer_token)
        self.workbook = None

    def open_file(self) -> openpyxl.Workbook:
        return openpyxl.load_workbook(self.file_path, keep_vba=False, data_only=True)
    
    def get_queries(self) -> list[TweetQuery]:
        self.workbook = self.open_file()
        queries = []

        if "input" in self.workbook.sheetnames:
            sheet = self.workbook["input"]
        else:
            sheet = self.workbook[self.workbook.sheetnames[0]]
        
        for row in sheet.iter_rows(min_row=2, max_row=2):
            queries.append(TweetQuery
            (row[0].value, row[1].value, row[2].value, row[3].value))
        
        return queries
    
    def pull_queries(self, queries: list[TweetQuery]) -> list[dict[TweetQueryHandler, list[TweetContainer]]]:
        queries_result = []
        for query in queries:
            handler = TweetQueryHandler(self.twitter_client, query)
            queries_result.append( {
                'handler' : handler,
                'result' : handler.pull()
            })
        return queries_result


t = TweetExtractor('AAAAAAAAAAAAAAAAAAAAAFV5UwEAAAAA97TibistsvE7HBSwwfCF97DMz%2FI%3DycD3aENBtw5YtaOoBxzM1h9lQJ4QgbJIOz9Fcj2TRIzIWNB95F', "./input.xlsx")
t.pull_queries(t.get_queries())







