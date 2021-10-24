from pysentimiento import SentimentAnalyzer
from TweetQueryHandler import TweetQueryHandler
from TweetContainer import TweetContainer
import openpyxl
from tqdm import tqdm
class TweetSentimentAnalyzer:
    
    def __init__(self, lang: str = "es") -> None:
        self.analizer = self.get_analizer(lang)

    def get_analizer(self, lang: str) -> SentimentAnalyzer.predict:
        if lang in ["es", "en"]:
            return SentimentAnalyzer(lang).predict
        else:
            return False
            #TODO ADD OTHER LANGS

    def from_list(
            self,
            wb: openpyxl.Workbook,  
            queries: list[dict[TweetQueryHandler, list[TweetContainer]]]
        ):
        results_sheet = wb.create_sheet("output")
        results_sheet.append(("query", "number_of_tweets", "positive", "neutral", "negative"))
        for query in tqdm(queries, "Analizando queries"):
            results = {
                "number_of_tweets": len(query["result"]),
                "positive": 0,
                "neutral": 0,
                "negative": 0
            }
            trans = {
                "POS": "positive",
                "NEU": "neutral",
                "NEG": "negative"
            }
            title = query["handler"].query.query
            sheet = None
            if title in wb.sheetnames:
                sheet = wb[title]
            else:
                sheet = wb.create_sheet(title)
                sheet.append(
                        (
                            "tweet_id",
                            "lang",
                            "date",
                            "text",
                            "filtered_text",
                            "eval_pos",
                            "eval_neu",
                            "eval_neg",
                            "result"
                        )
                    )
            for tweet in tqdm(query["result"], "Analizando tweets"):
                if(isinstance(tweet, TweetContainer)):
                    evaluation = self.analizer(tweet.filtered_text())
                    result_type = trans[evaluation.output]
                    results[result_type] = results[result_type] + 1
                    sheet.append(
                        tweet.get_row() + (evaluation.probas["POS"], evaluation.probas["NEU"], evaluation.probas["NEG"], evaluation.output)
                    )
            results_sheet.append((title, ) + tuple(results.values()))
        wb.save("output.xlsx")




